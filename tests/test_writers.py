import json
from pathlib import Path

from src.write_md import write_md
from src.write_json import write_json
from src.write_xml import write_xml
from src.write_conf import write_conf
from src.write_pdf import write_pdf
from src.write_epdf import write_epdf
from src.write_docx import write_docx
from src.write_xlsx import write_xlsx
from src.write_pptx import write_pptx
from src.writers import get_writer


def test_write_md_creates_file_with_correct_content(tmp_path: Path):
    target = tmp_path / "test.md"
    write_md("# Hello\nWorld", target)
    assert target.read_text() == "# Hello\nWorld"


def test_write_md_creates_parent_directories(tmp_path: Path):
    target = tmp_path / "deep" / "nested" / "file.md"
    write_md("content", target)
    assert target.exists()
    assert target.read_text() == "content"


def test_write_json_creates_valid_json(tmp_path: Path):
    target = tmp_path / "data.json"
    write_json('{"key": "value", "count": 42}', target)
    with open(target) as f:
        data = json.load(f)
    assert data == {"key": "value", "count": 42}


def test_write_xml_preserves_content(tmp_path: Path):
    target = tmp_path / "data.xml"
    content = "<root><item>data</item></root>"
    write_xml(content, target)
    assert target.read_text() == content


def test_write_conf_preserves_content(tmp_path: Path):
    target = tmp_path / "app.conf"
    content = "[section]\nkey=value\nother=123"
    write_conf(content, target)
    assert target.read_text() == content


def test_write_pdf_creates_valid_pdf(tmp_path: Path):
    target = tmp_path / "report.pdf"
    write_pdf("Some report content about quarterly results", target)
    raw = target.read_bytes()
    assert raw[:4] == b"%PDF"
    assert target.stat().st_size > 0


def test_write_epdf_creates_encrypted_pdf(tmp_path: Path):
    target = tmp_path / "secret.pdf"
    write_epdf("Secret content", target, password="abc123")
    assert target.exists()
    from pypdf import PdfReader
    reader = PdfReader(target, password="abc123")
    assert len(reader.pages) >= 1


def test_write_epdf_requires_password_to_read(tmp_path: Path):
    import pytest
    from pypdf import PdfReader
    target = tmp_path / "locked.pdf"
    write_epdf("Locked content", target, password="mypass")
    reader = PdfReader(target)
    with pytest.raises(Exception):
        reader.pages[0].extract_text()


def test_write_docx_creates_valid_docx(tmp_path: Path):
    from docx import Document
    target = tmp_path / "doc.docx"
    write_docx("# Heading\nParagraph text\n- bullet one\n- bullet two", target)
    assert target.exists()
    doc = Document(str(target))
    assert len(doc.paragraphs) > 0


def test_write_xlsx_creates_valid_spreadsheet(tmp_path: Path):
    from openpyxl import load_workbook
    target = tmp_path / "data.xlsx"
    write_xlsx("Name,Age\nAlice,30\nBob,25", target)
    assert target.exists()
    wb = load_workbook(target)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=1, column=2).value == "Age"
    assert ws.cell(row=2, column=1).value == "Alice"
    assert ws.cell(row=2, column=2).value == "30"


def test_write_pptx_creates_valid_presentation(tmp_path: Path):
    from pptx import Presentation
    target = tmp_path / "slides.pptx"
    write_pptx("Slide 1 Title\nBullet A\nBullet B\n---\nSlide 2 Title\nContent", target)
    assert target.exists()
    prs = Presentation(str(target))
    assert len(prs.slides) == 2


import pytest


def test_get_writer_returns_correct_functions():
    assert get_writer("md") == write_md
    assert get_writer("json") == write_json
    assert get_writer("xml") == write_xml
    assert get_writer("conf") == write_conf
    assert get_writer("pdf") == write_pdf
    assert get_writer("docx") == write_docx
    assert get_writer("xlsx") == write_xlsx
    assert get_writer("pptx") == write_pptx


def test_get_writer_epdf_returns_callable():
    writer = get_writer("epdf")
    assert callable(writer)


def test_get_writer_unknown_format_raises():
    with pytest.raises(ValueError):
        get_writer("exe")
